import openpyxl
from openpyxl.styles import Font

# Hostname, Last Seen, AV Status, Backup Status

path = 'C:\\Users\\athomatis\\SANTA\\Santa.xlsx'

workbook = openpyxl.load_workbook(path)
sheets = workbook.sheetnames

class SANTA():

    def __init__(self, sheet):

        self.worksheet = workbook[sheet]
        self.device_list = []
        self.user_exceptions_list = []
        self.av_exceptions_list = []
        self.backup_exceptions_list = []
        self.headers = ['Hostname', 'Last User', 'Last Online', 'AV Status', 'Backup Status', '', 'User Exceptions', 'AV Exceptions', 'Backup Exceptions']

    def get_devices(self):

        column = self.worksheet['A']

        for cell in column:
            value = str(cell.value)
            self.device_list.append(value)

    def get_users(self):

        # Get list of devices
        datto_list = []
        column = self.worksheet['B']

        for cell in column:
            value = str(cell.value)
            raw = value
            split = value.split('\\')[-1]
            datto_list.append({"raw" : raw, "split" : split.upper()})


        # Get list of last logged in users
        user_list = []
        column = self.worksheet['F']

        for cell in column:
            value = str(cell.value)
            raw = value
            split = value.split('@')[0]

            if raw == 'None':
                break

            user_list.append({"raw" : raw, "split" : split.upper()})


        return datto_list, user_list

    def get_anti_virus(self):

        # AV Devices List
        device_list = []
        column = self.worksheet['G']

        for cell in column:
            value = str(cell.value)
            value = value.split('.')[0]
            device_list.append(value)


        # AV Last Seen List
        anti_virus_list = []
        column = self.worksheet['H']

        for cell in column:
            value = str(cell.value)
            date = value.split(" ")

            if len(date) > 4:
                anti_virus_list.append(date[0])

            else:
                anti_virus_list.append(value)
        
        # List 3
        av_list = []

        for x in range(len(device_list)):

            if device_list[x] != 'None':

                av_list.append({'device': device_list[x], 'last_seen': anti_virus_list[x]})

        return av_list

    def get_backup(self):

        # Backup Devices List
        device_list = []
        column = self.worksheet['I']

        for cell in column:
            value = str(cell.value)
            device_list.append(value)


        # Last Backup List
        backup_list = []
        column = self.worksheet['J']

        for cell in column:
            value = str(cell.value)

            if value == '-':
                value = "No Data"

            date = value.split(" ")
            
            if len(date) > 3:
                backup_list.append(date[0])
            
            else:
                backup_list.append(value)
        
        # List 3
        crashplan_list = []

        for x in range(len(device_list)):

            if device_list[x] != 'None':

                crashplan_list.append({'device': device_list[x], 'last_backup': backup_list[x]})

        return crashplan_list

    def update_sheet(self, row, value, column):

        if value != 'None':
            self.worksheet.cell(row=row, column=column).value = value
        else:
            self.worksheet.cell(row=row, column=column).value = "/////////////////////////////////////////////////////////////////"

        workbook.save(path)

    def users_update(self):

        datto_list, user_list = self.get_users()
        
        # Exceptions List

        for x in user_list:

            self.user_exceptions_list.append(x['raw'])


        row = 0
        for x in datto_list:
            row += 1

            for y in user_list:

                if x['split'] == y['split']:
                    
                    if y['raw'] in self.user_exceptions_list:
                        self.user_exceptions_list.remove(y['raw'])

                    self.update_sheet(row, y['raw'], 2)
                    break

    def anti_virus_update(self):

        webroot_list = self.get_anti_virus()
        self.find_anti_virus(webroot_list)

        # Exceptions List

        for x in webroot_list:

            self.av_exceptions_list.append(x)

        row = 0
        for x in self.device_list:
            row += 1

            for y in webroot_list:

                if x.upper() == y['device'].upper():

                    for i in self.av_exceptions_list:

                        if i['device'] == y['device']:

                            self.av_exceptions_list.remove(i)
                            break

                    self.update_sheet(row, y['last_seen'], 3)
                    break

    def backup_update(self):

        backup_list = self.get_backup()

        # Exceptions List
        for x in backup_list:

            self.backup_exceptions_list.append(x)

        row = 0
        for x in self.device_list:
            row += 1

            for y in backup_list:

                if x.upper() == y['device'].upper():

                    for i in self.backup_exceptions_list:

                        if i['device'] == y['device']:

                            self.backup_exceptions_list.remove(i)
                            break

                    self.update_sheet(row, y['last_backup'], 4)
                    break

    def run(self):

        self.worksheet.insert_cols(4)

        self.get_devices()
        self.users_update()
        self.anti_virus_update()
        self.backup_update()
        self.finishing_touches()
        self.exceptions()

    def finishing_touches(self):

        # Delete Columns
        for x in range(5):
            self.worksheet.delete_cols(6)

        # Headers
        self.worksheet.insert_rows(1)
        for x in range(len(self.headers)):

            cell = self.worksheet.cell(row=1, column=x+1)
            cell.value = self.headers[x]
            cell.font = Font(name="Arial", size=12, bold=True)

        # Update Last Seen
        column = self.worksheet['C']
        for cell in column:

            value = str(cell.value)
            value = value.split("T")[0]
            cell.value = value


        workbook.save(path)

    def exceptions(self):

        # User Exceptions
        row = 1
        for x in self.user_exceptions_list:
            row += 1
            self.update_sheet(row, x, 6)

        # Antivirus Exceptions
        row = 1
        for x in self.av_exceptions_list:
            row += 1
            value = f"{x['device']} - {x['last_seen']}"
            self.update_sheet(row, value, 7)
        
        # Backup Exceptions 
        row = 1
        for x in self.backup_exceptions_list:
            row += 1
            value = f"{x['device']} - {x['last_backup']}"
            self.update_sheet(row, value, 8)

    def find_anti_virus(self, av_list):

        for x in av_list:

            if '.local' in x['device']:

                # Capture Client
                self.headers[3] = "Capture Client Status"
                self.headers[7] = "Capture Client Exceptions"
                return

        # Webroot
        self.headers[3] = "Webroot Status"
        self.headers[7] = "Webroot Exceptions"

for s in sheets:

    santa = SANTA(s)
    santa.run()

workbook.close()
